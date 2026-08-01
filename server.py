from dotenv import load_dotenv
from pathlib import Path
from fastapi_mail import MessageSchema, FastMail, ConnectionConfig
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import io
import csv
import uuid
import logging
import bcrypt
import jwt
import httpx
import openpyxl
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
conf = ConnectionConfig(
    MAIL_USERNAME=os.getenv("EMAIL"),
    MAIL_PASSWORD=os.getenv("PASSWORD"),
    MAIL_FROM=os.getenv("EMAIL"),
    MAIL_PORT=465,
    MAIL_SERVER="smtp.gmail.com",
    MAIL_SSL_TLS=True,
    MAIL_STARTTLS=False,
    USE_CREDENTIALS=True,
)

# ---- DB ----
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = "HS256"
ACCESS_TTL = timedelta(hours=12)
REFRESH_TTL = timedelta(days=7)

app = FastAPI(title="Student Event Evaluation Management System")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ---- helpers ----
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False


def create_token(sub: str, email: str, role: str, ttl: timedelta, ttype: str) -> str:
    payload = {
        "sub": sub, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + ttl, "type": ttype,
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def clean_doc(d: dict) -> dict:
    if not d:
        return d
    d.pop("_id", None)
    return d


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_role(*roles: str):
    async def _dep(user: dict = Depends(get_current_user)) -> dict:
        if user.get("role") not in roles:
            raise HTTPException(status_code=403, detail="Forbidden")
        return user
    return _dep


def set_auth_cookies(response: Response, access: str, refresh: str):
    response.set_cookie("access_token", access, httponly=True, secure=True, samesite="none", max_age=int(ACCESS_TTL.total_seconds()), path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="none", max_age=int(REFRESH_TTL.total_seconds()), path="/")

async def generate_team_id():
    last_team = await db.teams.find_one(
        {},
        sort=[("team_id", -1)]
    )

    if not last_team:
        return "T001"

    last_number = int(last_team["team_id"][1:])
    return f"T{last_number + 1:03d}"

async def student_already_in_team(
    student_id: str,
    event_id: str,
    ignore_team_id: str = None
):

    query = {
        "event_id": event_id,
        "member_ids": student_id
    }

    if ignore_team_id:
        query["id"] = {
            "$ne": ignore_team_id
        }

    return await db.teams.find_one(query)


# ---- Models ----
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class ChangePwIn(BaseModel):
    current_password: str
    new_password: str


class ProfileIn(BaseModel):
    name: Optional[str] = None
    department: Optional[str] = None
    designation: Optional[str] = None


class EventIn(BaseModel):
    event_name: str
    description: Optional[str] = ""
    event_type: Literal["Individual", "Group"]
    event_date: str
    event_time: str
    venue: str
    status: Literal["active", "completed", "draft"] = "active"


class ParameterIn(BaseModel):
    parameter_name: str
    weightage: float


class StudentIn(BaseModel):
    enrollment_no: str
    student_name: str
    email: str
    department: str
    semester: str
    institute: str
    phone_no: str
    whatsapp_no: str
    event_id: str
    status: Optional[Literal["active", "inactive"]] = "active"


class StudentStatusIn(BaseModel):
    status: Literal["active", "inactive"]


class EvaluatorIn(BaseModel):
    name: str
    email: EmailStr
    password: Optional[str] = "Eval@12345"
    department: str
    designation: str


class AssignmentIn(BaseModel):
    event_id: str
    evaluator_id: str
    student_ids: List[str]


class SheetSyncIn(BaseModel):
    sheet_id: str
    gid: Optional[str] = "0"


class EvaluationMarkIn(BaseModel):
    parameter_id: str
    marks: float


class EvaluationIn(BaseModel):
    student_id: str
    event_id: str
    comments: Optional[str] = ""
    marks: List[EvaluationMarkIn]

class EmailRequest(BaseModel):
    user_id: str
    event_id: str

class StudentDeleteRequest(BaseModel):
    event_id: str

class TeamIn(BaseModel):
    team_name: str
    event_id: str
    leader_id: str
    member_ids: List[str]
    status: Optional[Literal["active", "inactive"]] = "active"


class TeamStatusIn(BaseModel):
    status: Literal["active", "inactive"]


# ---- Auth Routes ----
app.include_router(api)
@app.get("/")
async def root():
    return {
        "status": "ok",
        "message": "Student Event Evaluation Management System API is running"
    }

@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if user.get("status") == "disabled":
        raise HTTPException(status_code=403, detail="Account disabled")
    access = create_token(user["id"], user["email"], user["role"], ACCESS_TTL, "access")
    refresh = create_token(user["id"], user["email"], user["role"], REFRESH_TTL, "refresh")
    set_auth_cookies(response, access, refresh)
    user.pop("_id", None); user.pop("password_hash", None)
    return {"user": user, "access_token": access}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


@api.post("/auth/refresh")
async def refresh(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        u = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not u:
            raise HTTPException(status_code=401, detail="User not found")
        access = create_token(u["id"], u["email"], u["role"], ACCESS_TTL, "access")
        set_auth_cookies(response, access, token)
        return {"ok": True}
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")


@api.post("/auth/change-password")
async def change_pw(body: ChangePwIn, user: dict = Depends(get_current_user)):
    u = await db.users.find_one({"id": user["id"]})
    if not verify_password(body.current_password, u["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": hash_password(body.new_password)}})
    return {"ok": True}


@api.put("/auth/profile")
async def update_profile(body: ProfileIn, user: dict = Depends(get_current_user)):
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if update:
        await db.users.update_one({"id": user["id"]}, {"$set": update})
        if user["role"] == "evaluator" and ("department" in update or "designation" in update):
            await db.evaluators.update_one({"user_id": user["id"]}, {"$set": {k: v for k, v in update.items() if k in ("department", "designation")}})
    return await db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0})


# ---- Events ----
@api.get("/events")
async def list_events(user: dict = Depends(get_current_user)):
    return await db.events.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@api.post("/events")
async def create_event(body: EventIn, user: dict = Depends(require_role("super_admin"))):
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()
    await db.events.insert_one(doc)
    return clean_doc(doc)


@api.get("/events/{event_id}")
async def get_event(event_id: str, user: dict = Depends(get_current_user)):
    e = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not e:
        raise HTTPException(404, "Event not found")
    return e


@api.put("/events/{event_id}")
async def update_event(event_id: str, body: EventIn, user: dict = Depends(require_role("super_admin"))):
    r = await db.events.update_one({"id": event_id}, {"$set": body.model_dump()})
    if r.matched_count == 0:
        raise HTTPException(404, "Event not found")
    return await db.events.find_one({"id": event_id}, {"_id": 0})


@api.delete("/events/{event_id}")
async def delete_event(event_id: str, user: dict = Depends(require_role("super_admin"))):
    await db.events.delete_one({"id": event_id})
    await db.event_parameters.delete_many({"event_id": event_id})
    await db.assignments.delete_many({"event_id": event_id})
    await db.evaluations.delete_many({"event_id": event_id})
    return {"ok": True}


# ---- Parameters ----
@api.get("/events/{event_id}/parameters")
async def list_parameters(event_id: str, user: dict = Depends(get_current_user)):
    return await db.event_parameters.find({"event_id": event_id}, {"_id": 0}).to_list(1000)


@api.post("/events/{event_id}/parameters")
async def add_parameter(event_id: str, body: ParameterIn, user: dict = Depends(require_role("super_admin"))):
    existing = await db.event_parameters.find({"event_id": event_id}).to_list(1000)
    total = sum(p["weightage"] for p in existing) + body.weightage
    if total > 100:
        raise HTTPException(400, f"Total weightage exceeds 100 (would be {total})")
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["event_id"] = event_id
    await db.event_parameters.insert_one(doc)
    return clean_doc(doc)


@api.put("/parameters/{param_id}")
async def update_parameter(param_id: str, body: ParameterIn, user: dict = Depends(require_role("super_admin"))):
    p = await db.event_parameters.find_one({"id": param_id})
    if not p:
        raise HTTPException(404, "Not found")
    others = await db.event_parameters.find({"event_id": p["event_id"], "id": {"$ne": param_id}}).to_list(1000)
    total = sum(x["weightage"] for x in others) + body.weightage
    if total > 100:
        raise HTTPException(400, f"Total weightage exceeds 100 (would be {total})")
    await db.event_parameters.update_one({"id": param_id}, {"$set": body.model_dump()})
    return await db.event_parameters.find_one({"id": param_id}, {"_id": 0})


@api.delete("/parameters/{param_id}")
async def delete_parameter(param_id: str, user: dict = Depends(require_role("super_admin"))):
    await db.event_parameters.delete_one({"id": param_id})
    return {"ok": True}


# ---- Students ----
@api.get("/students")
async def list_students(
    search: Optional[str] = None,
    event_id: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    q = {}

    if event_id:
        q["event_ids"] = event_id

    if search:
        q["$or"] = [
            {"enrollment_no": {"$regex": search, "$options": "i"}},
            {"student_name": {"$regex": search, "$options": "i"}},
            {"department": {"$regex": search, "$options": "i"}},
            {"institute": {"$regex": search, "$options": "i"}},
            {"phone_no": {"$regex": search, "$options": "i"}},
            {"whatsapp_no": {"$regex": search, "$options": "i"}},
        ]

    return await db.students.find(q, {"_id": 0}).to_list(2000)


@api.post("/students")
async def create_student(body: StudentIn, user: dict = Depends(require_role("super_admin"))):
    existing = await db.students.find_one({
        "enrollment_no": body.enrollment_no
    })

    if existing:

        if body.event_id in existing.get("event_ids", []):
            raise HTTPException(
                400,
                "Student is already registered for this event"
            )

        await db.students.update_one(
            {"_id": existing["_id"]},
            {
                "$addToSet": {
                    "event_ids": body.event_id
                }
            }
        )

        return {"message": "Student added to event successfully"}
    doc = body.model_dump()

    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = now_iso()

    doc["event_ids"] = [body.event_id]

    await db.students.insert_one(doc)

    return clean_doc(doc)


@api.put("/students/{sid}")
async def update_student(sid: str, body: StudentIn, user: dict = Depends(require_role("super_admin"))):
    r = await db.students.update_one({"id": sid}, {"$set": body.model_dump()})
    if r.matched_count == 0:
        raise HTTPException(404, "Not found")
    return await db.students.find_one({"id": sid}, {"_id": 0})

@api.delete("/students/{sid}")
async def delete_student(
    sid: str,
    body: StudentDeleteRequest,
    user: dict = Depends(require_role("super_admin"))
):
    student = await db.students.find_one({"id": sid})

    if not student:
        raise HTTPException(404, "Student not found")
    # Get all evaluations of the student
    evaluations = await db.evaluations.find(
        {
            "student_id": sid,
            "event_id": body.event_id
        },
        {
            "id": 1
        }
    ).to_list(None)

    evaluation_ids = [e["id"] for e in evaluations]

    # Delete all evaluation marks
    if evaluation_ids:
        await db.evaluation_marks.delete_many({
            "evaluation_id": {"$in": evaluation_ids}
        })

    # Delete evaluations
    await db.evaluations.delete_many({
        "student_id": sid,
        "event_id": body.event_id
    })

    # Delete assignments
    await db.assignments.delete_many({
        "student_id": sid,
        "event_id": body.event_id
    })

    # Delete student
    await db.students.update_one(
        {"id": sid},
        {
            "$pull": {
                "event_ids": body.event_id
            }
        }
    )

    student = await db.students.find_one({"id": sid})
    if not student.get("event_ids"):
        await db.students.delete_one({"id": sid})

    return {"ok": True}


@api.patch("/students/{sid}/status")
async def set_student_status(sid: str, body: StudentStatusIn, user: dict = Depends(require_role("super_admin"))):
    r = await db.students.update_one({"id": sid}, {"$set": {"status": body.status}})
    if r.matched_count == 0:
        raise HTTPException(404, "Not found")
    return await db.students.find_one({"id": sid}, {"_id": 0})


@api.get("/students/export")
async def students_export(user: dict = Depends(require_role("super_admin"))):

    students = await db.students.find().to_list(length=None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append([
        "enrollment_no",
        "student_name",
        "department",
        "semester",
        "institute",
        "email",
        "phone_no",
        "whatsapp_no",
        "status"
    ])

    for student in students:
        ws.append([
            student.get("enrollment_no", ""),
            student.get("student_name", ""),
            student.get("department", ""),
            student.get("semester", ""),
            student.get("institute", ""),
            student.get("email", ""),
            student.get("phone_no", ""),
            student.get("whatsapp_no", ""),
            student.get("status", "")
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=students.xlsx"
        }
    )


@api.post("/students/upload")
async def upload_students(file: UploadFile = File(...),event_id: str = Form(...),user: dict = Depends(require_role("super_admin"))):
    content = await file.read()
    print("Selected Event:", event_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")

    def normalize_header(header):
        return (
            str(header)
            .strip()
            .lower()
            .replace("-", "_")
            .replace(" ", "_")
        )

    header_map = {
        "name": "student_name",
        "student_name": "student_name",
        "enrollment_no": "enrollment_no",
        "phone": "phone_no",
        "phone_no": "phone_no",
        "whatsapp": "whatsapp_no",
        "whatsapp_no": "whatsapp_no",
        "whatsapp_number": "whatsapp_no",
        "phone_number": "phone_no",
        "enrollment": "enrollment_no",
    }

    headers = [
        header_map.get(normalize_header(h), normalize_header(h))
        for h in rows[0]
    ]
    required = ["enrollment_no", "student_name", "department", "semester", "institute", "email", "phone_no", "whatsapp_no"]
    for r in required:
        if r not in headers:
            raise HTTPException(400, f"Missing column: {r}")
    created = 0; skipped = 0
    for row in rows[1:]:
        if not any(row): continue
        d = {headers[i]: ("" if row[i] is None else str(row[i]).strip()) for i in range(len(headers))}
        if not d.get("enrollment_no"):
            skipped += 1; continue
        existing = await db.students.find_one(
            {"enrollment_no": d["enrollment_no"]}
        )

        if existing:
            # Already has this event?
            if event_id in existing.get("event_ids", []):
                skipped += 1
                continue

            # Add the new event
            await db.students.update_one(
                {"_id": existing["_id"]},
                {"$push": {"event_ids": event_id}}
            )

            created += 1
            continue
        d["id"] = str(uuid.uuid4()); d["created_at"] = now_iso(); d.setdefault("status", "active"); d["event_ids"] = [event_id]
        await db.students.insert_one(d)
        created += 1
    return {"created": created, "skipped": skipped}


# ---- Teams ----
@api.post("/teams")
async def create_team(
    body: TeamIn,
    user: dict = Depends(require_role("super_admin"))
):
    # Event validation
    event = await db.events.find_one({"id": body.event_id})
    if not event:
        raise HTTPException(404, "Event not found")

    # Duplicate team name validation
    existing_team = await db.teams.find_one({
        "team_name": body.team_name,
        "event_id": body.event_id
    })

    if existing_team:
        raise HTTPException(
            400,
            "Team name already exists for this event."
        )

    # Leader validation
    leader = await db.students.find_one({"id": body.leader_id})
    if not leader:
        raise HTTPException(404, "Leader not found")

    member_ids = list(set(body.member_ids))

    if body.leader_id not in member_ids:
        raise HTTPException(
            400,
            "Leader must be included in member_ids."
        )

    # Validate every member
    for sid in member_ids:

        student = await db.students.find_one({"id": sid})

        if not student:
            raise HTTPException(
                404,
                f"Student {sid} not found."
            )

        # Student cannot belong to another team of same event
        existing = await db.teams.find_one({
            "event_id": body.event_id,
            "member_ids": sid
        })

        if existing:
            raise HTTPException(
                400,
                f"{student['student_name']} already belongs to another team."
            )

    doc = {
        "id": str(uuid.uuid4()),
        "team_id": await generate_team_id(),
        "team_name": body.team_name,
        "event_id": body.event_id,
        "leader_id": body.leader_id,
        "member_ids": member_ids,
        "status": body.status,
        "created_at": now_iso()
    }

    await db.teams.insert_one(doc)

    return clean_doc(doc)

@api.get("/teams")
async def list_teams(
    search: Optional[str] = None,
    event_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 10,
    user: dict = Depends(get_current_user)
):

    query = {}

    if event_id:
        query["event_id"] = event_id

    if status:
        query["status"] = status

    total = await db.teams.count_documents(query)

    skip = (page - 1) * limit

    teams = await db.teams.find(
        query,
        {"_id": 0}
    ).skip(skip).limit(limit).to_list(limit)

    data = []

    for team in teams:

        leader = await db.students.find_one(
            {"id": team["leader_id"]},
            {"_id": 0}
        )

        event = await db.events.find_one(
            {"id": team["event_id"]},
            {"_id": 0}
        )

        members = await db.students.find(
            {
                "id": {
                    "$in": team["member_ids"]
                }
            },
            {"_id": 0}
        ).to_list(100)

        data.append({

            "id": team["id"],

            "team_id": team["team_id"],

            "team_name": team["team_name"],

            "leader_id": team["leader_id"],

            "event_id": team["event_id"],

            "status": team["status"],

            "leader": leader,

            "event": event,

            "member_count": len(members),

            "members": members

        })

    if search:

        search = search.lower()

        filtered = []

        for team in data:

            found = False

            if search in team["team_name"].lower():
                found = True

            elif search in team["team_id"].lower():
                found = True

            elif team["leader"] and search in team["leader"]["student_name"].lower():
                found = True

            else:

                for m in team["members"]:

                    if (
                        search in m.get("student_name","").lower()
                        or search in m.get("enrollment_no","").lower()
                    ):
                        found = True
                        break

            if found:
                filtered.append(team)

        data = filtered

    return {
        "data": data,
        "total": len(data)
    }

@api.get("/teams/{team_id}/members")
async def get_team_members(
    team_id: str,
    user: dict = Depends(get_current_user)
):

    team = await db.teams.find_one(
        {"id": team_id},
        {"_id": 0}
    )

    if not team:
        raise HTTPException(
            404,
            "Team not found"
        )

    members = await db.students.find(
        {
            "id": {
                "$in": team["member_ids"]
            }
        },
        {"_id": 0}
    ).to_list(100)

    return members

@api.put("/teams/{team_id}")
async def update_team(
    team_id: str,
    body: TeamIn,
    user: dict = Depends(require_role("super_admin"))
):

    team = await db.teams.find_one({"id": team_id})

    if not team:
        raise HTTPException(404, "Team not found")

    # Validate event
    event = await db.events.find_one({"id": body.event_id})

    if not event:
        raise HTTPException(404, "Event not found")

    # Validate leader
    leader = await db.students.find_one({"id": body.leader_id})

    if not leader:
        raise HTTPException(404, "Leader not found")

    member_ids = list(set(body.member_ids))

    if body.leader_id not in member_ids:
        raise HTTPException(
            400,
            "Leader must be included in member_ids."
        )

    # Duplicate Team Name
    duplicate = await db.teams.find_one({
        "team_name": body.team_name,
        "event_id": body.event_id,
        "id": {"$ne": team_id}
    })

    if duplicate:
        raise HTTPException(
            400,
            "Another team with this name already exists."
        )

    # Validate students
    for sid in member_ids:

        student = await db.students.find_one({"id": sid})

        if not student:
            raise HTTPException(
                404,
                f"Student {sid} not found."
            )

        existing = await db.teams.find_one({

            "event_id": body.event_id,

            "member_ids": sid,

            "id": {"$ne": team_id}

        })

        if existing:

            raise HTTPException(

                400,

                f"{student['student_name']} already belongs to another team."

            )

    await db.teams.update_one(

        {"id": team_id},

        {
            "$set": {

                "team_name": body.team_name,

                "event_id": body.event_id,

                "leader_id": body.leader_id,

                "member_ids": member_ids,

                "status": body.status

            }

        }

    )

    return await db.teams.find_one(
        {"id": team_id},
        {"_id": 0}
    )

@api.delete("/teams/{team_id}")
async def delete_team(
    team_id: str,
    user: dict = Depends(require_role("super_admin"))
):

    team = await db.teams.find_one({"id": team_id})

    if not team:
        raise HTTPException(
            404,
            "Team not found"
        )

    evaluation = await db.evaluations.find_one({
        "team_id": team_id
    })

    if evaluation:
        raise HTTPException(
            400,
            "This team has already been evaluated and cannot be deleted."
        )

    await db.teams.delete_one({
        "id": team_id
    })

    return {
        "ok": True
    }

@api.patch("/teams/{team_id}/status")
async def set_team_status(
    team_id: str,
    body: TeamStatusIn,
    user: dict = Depends(require_role("super_admin"))
):

    team = await db.teams.find_one({"id": team_id})

    if not team:
        raise HTTPException(
            404,
            "Team not found"
        )

    await db.teams.update_one(

        {"id": team_id},

        {
            "$set": {
                "status": body.status
            }
        }

    )

    return await db.teams.find_one(
        {"id": team_id},
        {"_id": 0}
    )

@api.post("/teams/upload")
async def upload_teams(
    event_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(require_role("super_admin"))
):

    event = await db.events.find_one({"id": event_id})

    if not event:
        raise HTTPException(404, "Event not found")

    wb = load_workbook(file.file)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(400, "Excel file is empty")

    headers = [
        str(h).strip() if h else ""
        for h in rows[0]
    ]

    # Required columns
    required = [
        "Team Name",
        "Leader Enrollment"
    ]

    for col in required:
        if col not in headers:
            raise HTTPException(
                400,
                f"Missing column: {col}"
            )

    idx_team = headers.index("Team Name")
    idx_leader = headers.index("Leader Enrollment")

    # Find all member enrollment columns
    member_columns = []

    for i, h in enumerate(headers):

        if "Member" in h and "Enrollment" in h:
            member_columns.append(i)

    inserted = 0

    for row in rows[1:]:

        if all(v is None for v in row):
            continue

        team_name = str(row[idx_team]).strip()

        leader_enrollment = str(
            row[idx_leader]
        ).strip()

        leader = await db.students.find_one({
            "enrollment_no": leader_enrollment
        })

        if not leader:
            raise HTTPException(
                400,
                f"Leader {leader_enrollment} not found."
            )

        # Duplicate Team
        duplicate = await db.teams.find_one({

            "team_name": team_name,

            "event_id": event_id

        })

        if duplicate:
            raise HTTPException(
                400,
                f"Team '{team_name}' already exists."
            )

        member_ids = [leader["id"]]

        for col in member_columns:

            if col >= len(row):
                continue

            value = row[col]

            if value is None:
                continue

            enrollment = str(value).strip()

            student = await db.students.find_one({
                "enrollment_no": enrollment
            })

            if not student:
                raise HTTPException(
                    400,
                    f"Student {enrollment} not found."
                )

            # Already in another team
            existing = await db.teams.find_one({
                "event_id": event_id,
                "member_ids": student["id"]
            })

            if existing:
                raise HTTPException(
                    400,
                    f"{student['student_name']} already belongs to another team."
                )

            member_ids.append(student["id"])

        member_ids = list(set(member_ids))

        await db.teams.insert_one({

            "id": str(uuid.uuid4()),

            "team_id": await generate_team_id(),

            "team_name": team_name,

            "event_id": event_id,

            "leader_id": leader["id"],

            "member_ids": member_ids,

            "status": "active",

            "created_at": now_iso()

        })

        inserted += 1

    return {

        "ok": True,

        "teams_created": inserted

    }

@api.get("/teams/export")
async def export_teams(
    event_id: str,
    user: dict = Depends(get_current_user)
):

    event = await db.events.find_one({"id": event_id})

    if not event:
        raise HTTPException(404, "Event not found")

    teams = await db.teams.find({
        "event_id": event_id
    }).to_list(1000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"

    headers = [

        "Team Name",

        "Leader Enrollment",
        "Leader Name",
        "Leader Email",
        "Leader Contact",
        "Leader Semester",
        "Leader Department",
        "Leader Institute",

        "Team Size",

        "Member1 Enrollment",
        "Member1 Name",

        "Member2 Enrollment",
        "Member2 Name",

        "Member3 Enrollment",
        "Member3 Name",

        "Member4 Enrollment",
        "Member4 Name",

        "Member5 Enrollment",
        "Member5 Name"

    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for team in teams:

        leader = await db.students.find_one({
            "id": team["leader_id"]
        })

        members = []

        for sid in team["member_ids"]:

            if sid == team["leader_id"]:
                continue

            student = await db.students.find_one({
                "id": sid
            })

            if student:
                members.append(student)

        row = [

            team["team_name"],

            leader.get("enrollment_no", ""),
            leader.get("student_name", ""),
            leader.get("email", ""),
            leader.get("phone_no", ""),
            leader.get("semester", ""),
            leader.get("department", ""),
            leader.get("institute", ""),

            len(team["member_ids"])

        ]

        for student in members:

            row.extend([

                student.get("enrollment_no", ""),

                student.get("student_name", "")

            ])

        while len(row) < len(headers):
            row.append("")

        ws.append(row)

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return StreamingResponse(

        output,

        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

        headers={

            "Content-Disposition":
            "attachment; filename=teams.xlsx"

        }

    )

# ---- Evaluators ----
@api.get("/evaluators")
async def list_evaluators(user: dict = Depends(require_role("super_admin"))):
    out = []
    async for e in db.evaluators.find({}, {"_id": 0}):
        u = await db.users.find_one({"id": e["user_id"]}, {"_id": 0, "password_hash": 0})
        if u:
            out.append({**e, "name": u["name"], "email": u["email"], "status": u.get("status", "active")})
    return out


@api.post("/evaluators")
async def create_evaluator(body: EvaluatorIn, user: dict = Depends(require_role("super_admin"))):
    em = body.email.lower().strip()
    if await db.users.find_one({"email": em}):
        raise HTTPException(400, "Email already exists")
    uid = str(uuid.uuid4())
    await db.users.insert_one({
        "id": uid, "name": body.name, "email": em,
        "password_hash": hash_password(body.password or "Eval@12345"),
        "role": "evaluator", "status": "active",
        "department": body.department, "designation": body.designation,
        "created_at": now_iso(),
    })
    eid = str(uuid.uuid4())
    await db.evaluators.insert_one({"id": eid, "user_id": uid, "department": body.department, "designation": body.designation})
    return {"id": eid, "user_id": uid, "name": body.name, "email": em, "department": body.department, "designation": body.designation}


@api.put("/evaluators/{eid}")
async def update_evaluator(eid: str, body: EvaluatorIn, user: dict = Depends(require_role("super_admin"))):
    e = await db.evaluators.find_one({"id": eid})
    if not e:
        raise HTTPException(404, "Not found")
    await db.evaluators.update_one({"id": eid}, {"$set": {"department": body.department, "designation": body.designation}})
    await db.users.update_one({"id": e["user_id"]}, {"$set": {"name": body.name, "department": body.department, "designation": body.designation}})
    return {"ok": True}


@api.delete("/evaluators/{eid}")
async def delete_evaluator(eid: str, user: dict = Depends(require_role("super_admin"))):
    e = await db.evaluators.find_one({"id": eid})
    if e:
        await db.users.delete_one({"id": e["user_id"]})
        await db.assignments.delete_many({"evaluator_id": eid})
    await db.evaluators.delete_one({"id": eid})
    return {"ok": True}


@api.get("/evaluators/export")
async def evaluators_export(
    user: dict = Depends(require_role("super_admin"))
):
    evaluators = await db.evaluators.find().to_list(length=None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluators"

    ws.append([
        "name",
        "email",
        "department",
        "designation"
    ])

    for evaluator in evaluators:

        user_data = await db.users.find_one({
            "id": evaluator.get("user_id")
        })

        ws.append([
            user_data.get("name", "") if user_data else "",
            user_data.get("email", "") if user_data else "",
            evaluator.get("department", ""),
            evaluator.get("designation", "")
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=evaluators.xlsx"
        }
    )


@api.post("/evaluators/upload")
async def upload_evaluators(file: UploadFile = File(...), user: dict = Depends(require_role("super_admin"))):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    created = 0; skipped = 0
    for row in rows[1:]:
        if not any(row): continue
        d = {headers[i]: ("" if row[i] is None else str(row[i]).strip()) for i in range(len(headers))}
        if not d.get("email"): skipped += 1; continue
        em = d["email"].lower()
        if await db.users.find_one({"email": em}):
            skipped += 1; continue
        uid = str(uuid.uuid4())
        await db.users.insert_one({
            "id": uid, "name": d.get("name", ""), "email": em,
            "password_hash": hash_password(d.get("password") or "Eval@12345"),
            "role": "evaluator", "status": "active",
            "department": d.get("department", ""), "designation": d.get("designation", ""),
            "created_at": now_iso(),
        })
        await db.evaluators.insert_one({"id": str(uuid.uuid4()), "user_id": uid,
                                         "department": d.get("department", ""), "designation": d.get("designation", "")})
        created += 1
    return {"created": created, "skipped": skipped}

@api.post("/send-email-evaluator")
async def send_email(data: EmailRequest):

    user = await db.users.find_one(
        {"id": data.user_id},
        {"_id": 0}
    )

    if not user:
        print(user)
        raise HTTPException(404, "User not found")

    event = await db.events.find_one(
        {"id": data.event_id},
        {"_id": 0}
    )

    if not event:
        raise HTTPException(404, "Event not found")

    message = MessageSchema(
        subject=f"Invitation to Serve as an Evaluator for {event['event_name']} | Parul University",
        recipients=[user["email"]],
        body=f"""
    Dear {user['name']},

    Greetings from the Technical Event Cell, Parul University.

    We are pleased to invite you to serve as an Evaluator for the upcoming technical event organized by the Technical Event Cell, Parul University. Your expertise and experience will be instrumental in ensuring a fair, transparent, and insightful evaluation of the participating teams.

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                EVENT DETAILS
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    Event Name : {event['event_name']}
    Date       : {event['event_date']}
    Time       : {event['event_time']}
    Venue      : {event['venue']}

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            EVALUATOR LOGIN
    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    User ID  : {user['email']}

    Password : YourFirstName@12345

    Example:
    If your name is "Shri Narendra Damodardas Modi",
    your password will be:

    Narendra@12345

    ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    Please log in to the Evaluation Portal using the above credentials. We recommend signing in at least 10–15 minutes before the event begins to ensure a smooth evaluation process.

    Your contribution will play a vital role in encouraging innovation, creativity, and technical excellence among our students. We sincerely appreciate your valuable time and support and look forward to your participation.

    If you have any questions or require any assistance, please do not hesitate to contact us.

    Thank you for accepting our invitation. We look forward to welcoming you to Parul University.

    Warm Regards,

    Technical Event Cell
    Parul University
    Vadodara, Gujarat

    Email   : technicaleventcell@paruluniversity.ac.in
    Contact : +91-XXXXXXXXXX
    """,
        subtype="plain",
    )

    fm = FastMail(conf)

    await fm.send_message(message)

    return {
        "message": f"Invitation sent successfully to {user['name']}"
    }


# ---- Assignments ----
@api.get("/assignments")
async def list_assignments(event_id: Optional[str] = None, evaluator_id: Optional[str] = None,
                            user: dict = Depends(get_current_user)):
    q = {}
    if event_id: q["event_id"] = event_id
    if evaluator_id: q["evaluator_id"] = evaluator_id
    return await db.assignments.find(q, {"_id": 0}).to_list(5000)


@api.post("/assignments")
async def create_assignments(body: AssignmentIn, user: dict = Depends(require_role("super_admin"))):
    created = 0
    for sid in body.student_ids:
        exists = await db.assignments.find_one({"event_id": body.event_id, "evaluator_id": body.evaluator_id, "student_id": sid})
        if exists: continue
        await db.assignments.insert_one({
            "id": str(uuid.uuid4()), "event_id": body.event_id,
            "evaluator_id": body.evaluator_id, "student_id": sid,
            "assigned_at": now_iso(),
        })
        created += 1
    return {"created": created}


@api.delete("/assignments/{aid}")
async def delete_assignment(aid: str, user: dict = Depends(require_role("super_admin"))):
    await db.assignments.delete_one({"id": aid})
    return {"ok": True}


# ---- Evaluator-specific ----
async def _evaluator_id_for_user(uid: str) -> Optional[str]:
    e = await db.evaluators.find_one({"user_id": uid})
    return e["id"] if e else None


@api.get("/evaluator/assigned")
async def evaluator_assigned(event_id: Optional[str] = None, user: dict = Depends(require_role("evaluator"))):
    eid = await _evaluator_id_for_user(user["id"])
    if not eid: return []
    q = {"evaluator_id": eid}
    if event_id: q["event_id"] = event_id
    rows = await db.assignments.find(q, {"_id": 0}).to_list(5000)
    out = []
    for a in rows:
        s = await db.students.find_one({"id": a["student_id"]}, {"_id": 0})
        ev = await db.events.find_one({"id": a["event_id"]}, {"_id": 0})
        done = await db.evaluations.find_one({"student_id": a["student_id"], "event_id": a["event_id"], "evaluator_id": eid})
        out.append({**a, "student": s, "event": ev, "completed": bool(done),
                    "evaluation_id": done["id"] if done else None})
    return out


@api.get("/evaluator/dashboard")
async def evaluator_dashboard(user: dict = Depends(require_role("evaluator"))):
    eid = await _evaluator_id_for_user(user["id"])
    if not eid:
        return {"assigned": 0, "completed": 0, "pending": 0}
    total = await db.assignments.count_documents({"evaluator_id": eid})
    completed = await db.evaluations.count_documents({"evaluator_id": eid})
    return {"assigned": total, "completed": completed, "pending": max(total - completed, 0)}


# ---- Evaluations ----
@api.post("/evaluations")
async def submit_evaluation(body: EvaluationIn, user: dict = Depends(require_role("evaluator"))):
    eid = await _evaluator_id_for_user(user["id"])
    if not eid:
        raise HTTPException(400, "Evaluator profile missing")
    lock = await db.final_submissions.find_one({
        "evaluator_id": eid,
        "locked": True
    })

    if lock:
        raise HTTPException(
            status_code=403,
            detail="Final submission already completed. Evaluations are locked."
        )
    params = {p["id"]: p for p in await db.event_parameters.find({"event_id": body.event_id}).to_list(100)}
    total = 0.0
    for m in body.marks:
        p = params.get(m.parameter_id)
        if not p:
            raise HTTPException(400, f"Invalid parameter {m.parameter_id}")
        if m.marks < 0 or m.marks > p["weightage"]:
            raise HTTPException(400, f"Marks for '{p['parameter_name']}' must be 0..{p['weightage']}")
        total += m.marks
    existing = await db.evaluations.find_one({"student_id": body.student_id, "event_id": body.event_id, "evaluator_id": eid})
    if existing:
        await db.evaluations.update_one({"id": existing["id"]}, {"$set": {"comments": body.comments, "total_marks": total, "submitted_at": now_iso()}})
        await db.evaluation_marks.delete_many({"evaluation_id": existing["id"]})
        evid = existing["id"]
    else:
        evid = str(uuid.uuid4())
        await db.evaluations.insert_one({
            "id": evid, "student_id": body.student_id, "event_id": body.event_id,
            "evaluator_id": eid, "comments": body.comments,
            "total_marks": total, "submitted_at": now_iso(),
        })
    for m in body.marks:
        await db.evaluation_marks.insert_one({"id": str(uuid.uuid4()), "evaluation_id": evid,
                                               "parameter_id": m.parameter_id, "marks": m.marks})
    return {"id": evid, "total_marks": total}


@api.get("/evaluations/{evid}")
async def get_evaluation(evid: str, user: dict = Depends(get_current_user)):
    ev = await db.evaluations.find_one({"id": evid}, {"_id": 0})
    if not ev:
        raise HTTPException(404, "Not found")
    marks = await db.evaluation_marks.find({"evaluation_id": evid}, {"_id": 0}).to_list(100)
    s = await db.students.find_one({"id": ev["student_id"]}, {"_id": 0})
    e = await db.events.find_one({"id": ev["event_id"]}, {"_id": 0})
    params = await db.event_parameters.find({"event_id": ev["event_id"]}, {"_id": 0}).to_list(100)
    return {"evaluation": ev, "marks": marks, "student": s, "event": e, "parameters": params}


@api.get("/evaluator/completed")
async def evaluator_completed(user: dict = Depends(require_role("evaluator"))):
    eid = await _evaluator_id_for_user(user["id"])
    if not eid: return []
    rows = await db.evaluations.find({"evaluator_id": eid}, {"_id": 0}).sort("submitted_at", -1).to_list(2000)
    out = []
    for ev in rows:
        s = await db.students.find_one({"id": ev["student_id"]}, {"_id": 0})
        e = await db.events.find_one({"id": ev["event_id"]}, {"_id": 0})
        out.append({**ev, "student": s, "event": e})
    return out
@api.get("/evaluator/final-status")
async def final_status(user=Depends(get_current_user)):
    eid = await _evaluator_id_for_user(user["id"])

    lock = await db.final_submissions.find_one(
        {"evaluator_id": eid}
    )

    return {"locked": bool(lock)}


@api.post("/evaluator/final-submit")
async def final_submit(user=Depends(get_current_user)):
    eid = await _evaluator_id_for_user(user["id"])

    await db.final_submissions.update_one(
        {"evaluator_id": eid},
        {
            "$set": {
                "evaluator_id": eid,
                "locked": True,
                "submitted_at": datetime.utcnow()
            }
        },
        upsert=True
    )

    return {"success": True}


# ---- Results ----
def grade_for(score_pct: float) -> str:
    if score_pct >= 90: return "A+"
    if score_pct >= 80: return "A"
    if score_pct >= 70: return "B+"
    if score_pct >= 60: return "B"
    if score_pct >= 50: return "C"
    if score_pct >= 40: return "D"
    return "F"


@api.get("/results")
async def results(event_id: Optional[str] = None, department: Optional[str] = None,
                  semester: Optional[str] = None, user: dict = Depends(require_role("super_admin"))):
    q = {}
    if event_id: q["event_id"] = event_id
    evals = await db.evaluations.find(q, {"_id": 0}).to_list(10000)
    by_student = {}
    for ev in evals:
        key = (ev["student_id"], ev["event_id"])
        by_student.setdefault(key, []).append(ev["total_marks"])
    out = []
    for (sid, eid), marks in by_student.items():
        s = await db.students.find_one({"id": sid}, {"_id": 0})
        if not s: continue
        if department and s.get("department") != department: continue
        if semester and str(s.get("semester")) != str(semester): continue
        e = await db.events.find_one({"id": eid}, {"_id": 0})
        avg = sum(marks) / len(marks)
        out.append({
            "student_id": sid, "event_id": eid, "enrollment_no": s["enrollment_no"],
            "student_name": s["student_name"], "department": s["department"],
            "semester": s["semester"], "event_name": e["event_name"] if e else "",
            "phone_no": s.get("phone_no", ""),
            "whatsapp_no": s.get("whatsapp_no", ""),
            "score": round(avg, 2), "grade": grade_for(avg), "evaluator_count": len(marks),
        })
    out.sort(key=lambda x: x["score"], reverse=True)
    for i, r in enumerate(out): r["rank"] = i + 1
    return out


# ---- Dashboard ----
@api.get("/dashboard/admin")
async def admin_dashboard(user: dict = Depends(require_role("super_admin"))):
    total_events = await db.events.count_documents({})
    total_students = await db.students.count_documents({})
    total_evaluators = await db.evaluators.count_documents({})
    total_assignments = await db.assignments.count_documents({})
    completed = await db.evaluations.count_documents({})
    events = await db.events.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    event_perf = []
    for e in events[:8]:
        c = await db.evaluations.count_documents({"event_id": e["id"]})
        a = await db.assignments.count_documents({"event_id": e["id"]})
        event_perf.append({"event_name": e["event_name"], "completed": c, "assigned": a})
    recent_events = events[:5]
    return {
        "total_events": total_events, "total_students": total_students,
        "total_evaluators": total_evaluators, "completed_evaluations": completed,
        "total_assignments": total_assignments,
        "event_performance": event_perf, "recent_events": recent_events,
        "completion": {"completed": completed, "pending": max(total_assignments - completed, 0)},
    }


# ---- Reports ----
def _xlsx_response(rows: List[List], sheet_name: str, filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = sheet_name
    for r in rows: ws.append(r)
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 10), 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _pdf_response(title: str, headers: List[str], rows: List[List], filename: str):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor("#0033A0"))
    elements = [Paragraph(title, h), Spacer(1, 12)]
    data = [headers] + rows
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0033A0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E4E7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F4F5")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@api.get("/reports/ranking")
async def report_ranking(event_id: str, fmt: str = "xlsx", user: dict = Depends(require_role("super_admin"))):
    data = await results(event_id=event_id, user=user)
    headers = ["Rank", "Enrollment", "Student", "Department", "Semester", "Phone No", "Whatsapp No", "Event", "Score", "Grade"]
    rows = [[r["rank"], r["enrollment_no"], r["student_name"], r["department"], r["semester"], r["phone_no"],r["whatsapp_no"], r["event_name"], r["score"], r["grade"]] for r in data]
    if fmt == "pdf":
        return _pdf_response("Ranking Report", headers, rows, "ranking_report.pdf")
    return _xlsx_response([headers] + rows, "Ranking", "ranking_report.xlsx")


@api.get("/reports/event-wise")
async def report_event_wise(fmt: str = "xlsx", user: dict = Depends(require_role("super_admin"))):
    events = await db.events.find({}, {"_id": 0}).to_list(1000)
    headers = ["Event", "Date", "Status", "Assignments", "Evaluations Completed"]
    rows = []
    for e in events:
        a = await db.assignments.count_documents({"event_id": e["id"]})
        c = await db.evaluations.count_documents({"event_id": e["id"]})
        rows.append([e["event_name"], e.get("event_date", ""), e.get("status", ""), a, c])
    if fmt == "pdf":
        return _pdf_response("Event-Wise Report", headers, rows, "event_report.pdf")
    return _xlsx_response([headers] + rows, "Events", "event_report.xlsx")


@api.get("/reports/student-wise")
async def report_student_wise(fmt: str = "xlsx", user: dict = Depends(require_role("super_admin"))):
    students = await db.students.find({}, {"_id": 0}).to_list(5000)
    headers = ["Enrollment", "Student", "Department", "Semester", "Phone No", "Whatsapp No", "Events Evaluated", "Avg Score"]
    rows = []
    for s in students:
        evs = await db.evaluations.find({"student_id": s["id"]}).to_list(500)
        if not evs:
            rows.append([s["enrollment_no"], s["student_name"], s["department"], s.get("semester", ""), s["phone_no"], s["whatsapp_no"], 0, "-"])
            continue
        avg = sum(e["total_marks"] for e in evs) / len(evs)
        rows.append([s["enrollment_no"], s["student_name"], s["department"], s.get("semester", ""), s["phone_no"], s["whatsapp_no"], len(evs), round(avg, 2)])
    if fmt == "pdf":
        return _pdf_response("Student-Wise Report", headers, rows, "student_report.pdf")
    return _xlsx_response([headers] + rows, "Students", "student_report.xlsx")


@api.get("/reports/evaluator-wise")
async def report_evaluator_wise(fmt: str = "xlsx", user: dict = Depends(require_role("super_admin"))):
    evals_users = await db.evaluators.find({}, {"_id": 0}).to_list(1000)
    headers = ["Evaluator", "Email", "Department", "Designation", "Assigned", "Completed", "Pending"]
    rows = []
    for e in evals_users:
        u = await db.users.find_one({"id": e["user_id"]}, {"_id": 0})
        if not u: continue
        a = await db.assignments.count_documents({"evaluator_id": e["id"]})
        c = await db.evaluations.count_documents({"evaluator_id": e["id"]})
        rows.append([u["name"], u["email"], e.get("department", ""), e.get("designation", ""), a, c, max(a - c, 0)])
    if fmt == "pdf":
        return _pdf_response("Evaluator-Wise Report", headers, rows, "evaluator_report.pdf")
    return _xlsx_response([headers] + rows, "Evaluators", "evaluator_report.xlsx")


# ---- Credentials Excel (admin + evaluators) ----
@api.get("/credentials/export")
async def credentials_export(user: dict = Depends(require_role("super_admin"))):
    """Download all login credentials as Excel. Passwords are shown as '(set)' for
    existing users — set a new password in the cell to update on re-upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credentials"
    ws.append(["role", "name", "email", "password", "department", "designation", "status"])
    users = await db.users.find({}, {"_id": 0}).to_list(2000)
    for u in users:
        ws.append([
            u.get("role", ""),
            u.get("name", ""),
            u.get("email", ""),
            "(set)",
            u.get("department", ""),
            u.get("designation", ""),
            u.get("status", "active"),
        ])
    # Adjust widths
    for col in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(ml + 2, 12), 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=credentials.xlsx"},
    )


@api.post("/credentials/import")
async def credentials_import(file: UploadFile = File(...),
                              user: dict = Depends(require_role("super_admin"))):
    """Bulk upsert users from the credentials Excel.
    - Row matched by email.
    - If password cell is empty or '(set)', existing hash is preserved.
    - role='evaluator' creates linked evaluator profile."""
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")
    headers = [str(h).strip() if h else "" for h in rows[0]]
    required = ["role", "name", "email"]
    for r in required:
        if r not in headers:
            raise HTTPException(400, f"Missing column: {r}")
    created = updated = skipped = 0
    for row in rows[1:]:
        if not any(row): continue
        d = {headers[i]: ("" if row[i] is None else str(row[i]).strip()) for i in range(len(headers))}
        em = d.get("email", "").lower()
        role = d.get("role", "").lower()
        if not em or role not in ("super_admin", "evaluator"):
            skipped += 1; continue
        pw = d.get("password", "")
        status = d.get("status", "active") or "active"
        update = {
            "name": d.get("name", ""),
            "role": role,
            "department": d.get("department", ""),
            "designation": d.get("designation", ""),
            "status": status,
        }
        existing = await db.users.find_one({"email": em})
        if existing:
            if pw and pw != "(set)":
                update["password_hash"] = hash_password(pw)
            await db.users.update_one({"email": em}, {"$set": update})
            if role == "evaluator":
                ev = await db.evaluators.find_one({"user_id": existing["id"]})
                ev_update = {"department": update["department"], "designation": update["designation"]}
                if ev:
                    await db.evaluators.update_one({"id": ev["id"]}, {"$set": ev_update})
                else:
                    await db.evaluators.insert_one({"id": str(uuid.uuid4()), "user_id": existing["id"], **ev_update})
            updated += 1
        else:
            if not pw or pw == "(set)":
                skipped += 1; continue
            uid = str(uuid.uuid4())
            await db.users.insert_one({
                "id": uid, "email": em, "password_hash": hash_password(pw),
                "created_at": now_iso(), **update,
            })
            if role == "evaluator":
                await db.evaluators.insert_one({"id": str(uuid.uuid4()), "user_id": uid,
                                                "department": update["department"], "designation": update["designation"]})
            created += 1
    return {"created": created, "updated": updated, "skipped": skipped}

@api.post("/credentials/sync-sheet")
async def credentials_sync_sheet(body: SheetSyncIn, user: dict = Depends(require_role("super_admin"))):
    """Pull credentials live from a public Google Sheet (Anyone with link can view).
    Expected header: role, name, email, password, department, designation, status."""
    url = f"https://docs.google.com/spreadsheets/d/{body.sheet_id}/export?format=csv&gid={body.gid or '0'}"
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=15.0) as cli:
            resp = await cli.get(url)
    except httpx.HTTPError as e:
        raise HTTPException(502, f"Could not reach Google Sheets: {e}")
    if resp.status_code != 200 or "text/csv" not in resp.headers.get("content-type", ""):
        raise HTTPException(
            400,
            "Sheet is not publicly accessible. In Google Sheets click Share → General access → "
            "'Anyone with the link' → Viewer, then try again.",
        )
    reader = csv.reader(io.StringIO(resp.text))
    rows = list(reader)
    if not rows:
        raise HTTPException(400, "Sheet is empty")
    headers = [h.strip().lower() for h in rows[0]]
    for r in ("role", "name", "email"):
        if r not in headers:
            raise HTTPException(400, f"Missing column: {r}")
    created = updated = skipped = 0
    for row in rows[1:]:
        if not any(c.strip() for c in row): continue
        d = {headers[i]: (row[i].strip() if i < len(row) and row[i] is not None else "") for i in range(len(headers))}
        em = d.get("email", "").lower()
        role = d.get("role", "").lower()
        if not em or role not in ("super_admin", "evaluator"):
            skipped += 1; continue
        pw = d.get("password", "")
        status = d.get("status", "active") or "active"
        update = {
            "name": d.get("name", ""), "role": role,
            "department": d.get("department", ""), "designation": d.get("designation", ""),
            "status": status,
        }
        existing = await db.users.find_one({"email": em})
        if existing:
            if pw and pw != "(set)":
                update["password_hash"] = hash_password(pw)
            await db.users.update_one({"email": em}, {"$set": update})
            if role == "evaluator":
                ev = await db.evaluators.find_one({"user_id": existing["id"]})
                ev_update = {"department": update["department"], "designation": update["designation"]}
                if ev:
                    await db.evaluators.update_one({"id": ev["id"]}, {"$set": ev_update})
                else:
                    await db.evaluators.insert_one({"id": str(uuid.uuid4()), "user_id": existing["id"], **ev_update})
            updated += 1
        else:
            if not pw or pw == "(set)":
                skipped += 1; continue
            uid = str(uuid.uuid4())
            await db.users.insert_one({
                "id": uid, "email": em, "password_hash": hash_password(pw),
                "created_at": now_iso(), **update,
            })
            if role == "evaluator":
                await db.evaluators.insert_one({"id": str(uuid.uuid4()), "user_id": uid,
                                                "department": update["department"], "designation": update["designation"]})
            created += 1
    return {"created": created, "updated": updated, "skipped": skipped, "rows_seen": len(rows) - 1}





@api.get("/lookup/student/{sid}")
async def lookup_student(sid: str, user: dict = Depends(get_current_user)):
    print("Received sid:", sid)

    s = await db.students.find_one({"id": sid}, {"_id": 0})

    print("Student:", s)

    if not s:
        raise HTTPException(404, "Student not found in DB")

    return s


# ---- Startup ----
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.events.create_index("id", unique=True)
    await db.students.create_index("enrollment_no", unique=True)
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@evalsystem.in").lower()
    admin_pw = os.environ.get("ADMIN_PASSWORD", "Admin@12345")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "name": "Super Admin", "email": admin_email,
            "password_hash": hash_password(admin_pw), "role": "super_admin",
            "status": "active", "created_at": now_iso(),
        })
        logger.info(f"Seeded admin: {admin_email}")
    elif not verify_password(admin_pw, existing.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_pw)}})
        logger.info(f"Updated admin password for {admin_email}")


@app.on_event("shutdown")
async def shutdown():
    client.close()


app.include_router(api)

_cors_env = os.environ.get("CORS_ORIGINS", "http://localhost:3000")
_origins = [o.strip() for o in _cors_env.split(",") if o.strip()]
logger.info(f"CORS Origins: {_origins}")
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)
